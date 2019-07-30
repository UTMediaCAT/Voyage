# Filename: articles_conversion_script.py
# Author: Raiyan Rahman
# Date: July 10th, 2019
# Description: This is a conversion script that converts an excel file to a
# JSON file containing the outlet names, websites, and type in the first,
# second, and third columns respectively.
# REQUIREMENT: Each section in the sheets must be separated from the next by
# 3 rows.

import json
from openpyxl import load_workbook


def convert_articles_to_json(filename: str) -> None:
    """
    Create a JSON file containing the converted data from the given excel
    file.
    The headers of the columns are stored as follows:
    Outlet Name, Website, Type
    :param filename: The name of the excel file to be converted.
    :return: None
    REQ: The filename must point to a valid .xlsx file with the required data.
    """
    headers = {
        1: 'Outlet Name',
        2: 'Website',
        3: 'Type'
    }
    # Open the excel file.
    wb = load_workbook(filename)
    # Create a dictionary and store the data for each sheet in it.
    excel_data = get_articles_excel_data(wb, headers)
    # Create the JSON and save it.
    dump_data = json.dumps(excel_data)
    outfile = open('articles.json', 'a')
    outfile.write(dump_data)
    outfile.close()
    return


def get_articles_excel_data(wb, headers: dict) -> dict:
    """
    Return the parsed excel data from the given workbook with respect to the
    given dictionary of headers.
    :param wb: The excel file's workbook.
    :param headers: The dictionary of headers.
    :return: The dictionary containing the parsed excel sheet data.
    """
    # Create the dictionary to hold all the excel data.
    parsed_data = {}
    # Get the sheet names.
    sheet_names = wb.sheetnames
    # Loop through the sheets and get sheet data to parsed data.
    for sheet in sheet_names:
        parsed_data[sheet] = get_articles_sheet_data(wb[sheet], headers)
    # Return the data.
    return parsed_data


def get_articles_sheet_data(sheet, headers: dict) -> dict:
    """
    Return the parsed data for a single excel sheet with respect to the given
    dictionary of headers.
    :param sheet: The worksheet.
    :param headers: The dictionary headers.
    :return: The dictionary containing the parsed sheet data.
    """
    sheet_data = {}
    curr_header = 'Unlabelled'
    empty_row_counter = 0
    curr_section = []
    # Loop through all the rows and columns in the sheet.
    for row in range(2, sheet.max_row):
        # Check if the row is a header, only the first col is nonempty.
        if sheet.cell(row=row, column=1).value is not None and \
                sheet.cell(row=row, column=2).value is None and \
                sheet.cell(row=row, column=3).value is None:
            # Start a new section.
            curr_header = sheet.cell(row=row, column=1).value.strip()
            curr_section = []
            empty_row_counter = 0
        # If row is empty, all columns are too.
        elif sheet.cell(row=row, column=1).value is None and \
                sheet.cell(row=row, column=2).value is None and \
                sheet.cell(row=row, column=3).value is None:
            # If the row is empty.
            if empty_row_counter is 2:
                # Add it to the section.
                sheet_data[curr_header] = curr_section
            elif empty_row_counter is 3:
                # End of all rows is reached.
                break
            # Increment how many empty rows were found.
            empty_row_counter += 1
        # If the row contains data.
        else:
            row_data = {}
            # Add the row data to the section data.
            for col in range(1, 4):
                if sheet.cell(row=row, column=col).value is not None:
                    row_data[headers[col]] = \
                        sheet.cell(row=row, column=col).value.strip()
            # Add the row data to the section.
            curr_section.append(row_data)
    # Return the sheet data.
    return sheet_data


if __name__ == '__main__':
    file_name = input('Enter filename of Excel File: ')
    convert_articles_to_json(filename=file_name)
