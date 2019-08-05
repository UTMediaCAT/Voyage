# Filename: twitter_handles_conversion_script.py
# Author: Raiyan Rahman
# Date: July 10th, 2019
# Description: This is a conversion script that converts an excel file to a
# JSON file containing the names, position, twitter handle in the first,
# second, and third column respectively. The file may also contain an optional
# authenticated column as its fourth column.

import json
from openpyxl import load_workbook


def convert_twitter_handles_to_json(filename: str) -> None:
    """
    Create a JSON file containing the converted data from the given excel
    file.
    The headers of the columns are stored as follows:
    Name, Position, Twitter Handle, (Optional Column), (Optional Notes Column)
    :param filename: The name of the excel file to be converted.
    :return: None
    REQ: The filename must point to a valid .xlsx file with the required data.
    """
    # Dictionary containing the headers.
    headers = {
        1: 'Name',
        2: 'Twitter Handle',
        3: 'Source/Referring',
        4: 'Domain',
        5: 'Position',
        6: 'Authenticated'
    }
    # Open the excel file.
    wb = load_workbook(filename)
    # Create a dictionary and store the data for each sheet in it.
    excel_data = get_twitter_excel_data(wb, headers)
    # Create the JSON and save it.
    dump_data = json.dumps(excel_data)
    outfile = open('twitter_handles.json', 'w')
    outfile.write(dump_data)
    outfile.close()
    return


def get_twitter_excel_data(wb, headers: dict) -> dict:
    """
    Return the parsed excel data from the given workbook with respect to the
    given dictionary of headers.
    :param wb: The excel file's workbook.
    :param headers: The dictionary of headers.
    :return: The dictionary containing the parsed excel sheets data.
    """
    # Create the dictionary to hold all the excel data.
    parsed_data = {}
    # Get the sheet names.
    sheet_names = wb.sheetnames
    # Loop through the sheets and get sheet data to parsed data.
    for sheet in sheet_names:
        parsed_data[sheet] = get_twitter_sheet_data(wb[sheet], headers)
    # Return the data.
    return parsed_data


def get_twitter_sheet_data(sheet, headers: dict) -> list:
    """
    Return the parsed data for a single excel sheet with respect to the given
    dictionary of headers.
    :param sheet: The worksheet.
    :param headers: The dictionary of headers.
    :return: The list containing the parsed sheet data.
    """
    sheet_data = []
    # Loop through the first row and update the dictionary of headers.
    done = False
    row = 1
    max_column = 1
    while not done:
        if sheet.cell(row=row, column=max_column).value != headers[max_column]:
            headers[max_column] = sheet.cell(row=row, column=max_column).value.strip()
        if sheet.cell(row=1, column=max_column).value is None or max_column is 6:
            done = True
        else:
            max_column += 1
    row += 1
    # Loop through the rows and get the data.
    done = False
    while not done:
        # If the next two rows are empty.
        if sheet.cell(row=row, column=1).value is None and \
                sheet.cell(row=row + 1, column=1).value is None:
            done = True
            pass
        # If this row is empty, skip it.
        elif sheet.cell(row=row, column=1).value is None:
            row += 1
            pass
        else:
            # Get all the row data.
            row_data = {}
            for i in range(1, max_column + 1):
                if sheet.cell(row=row, column=i).value is not None:
                    # Convert Yes/No to bool.
                    if headers[i] == 'Authenticated':
                        val = sheet.cell(row=row, column=i).value.strip()
                        val = val.lower()[0]
                        val = True if val == 'y' else False
                        row_data[headers[i]] = val
                    else:
                        # Normalize unicode to python string.
                        val = sheet.cell(row=row, column=i).value.strip()
                        if type(val) == 'unicode':
                            val = val.encode('ascii', 'ignore')
                        row_data[headers[i]] = val
                # Insert an empty string.
                else:
                    row_data[headers[i]] = ''
            # Add the row data and increment the row counter.
            sheet_data.append(row_data)
            row += 1
    # Return the data.
    return sheet_data


if __name__ == '__main__':
    file_name = input('Enter filename of Excel File: ')
    convert_twitter_handles_to_json(file_name)